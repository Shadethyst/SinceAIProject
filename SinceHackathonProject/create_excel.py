from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

example_data = [
    ["4-pole 125A switchfuse", 1, "Mittauskeskus MK1C", "AMCMK 4x95/29"],
    ["3-pole C16 circuiut breaker", 1, "Kerhotila liesi", "MMJ 5x2,5S"],
    ["1-pole C16 circuit breaker", 1, "Atk-ristikytkentä", "MMJ 3x2,5S"],
    ["3-pole C10 circuit breaker", 1, "", "?"],
    ["3-pole contractor", 1, "", ""],
    ["3-step switch", 1, "1-0-A", ""],
    ["Signal light", 1 ,"", ""]
]

def create_excel(name="Laskentatiedosto", data=example_data, missing=None):
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Laskentataulukko"
    
    headers = ["Product", "Amount", "Address", "Cable"]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="67A7DB", end_color="67A7DB", fill_type="solid")
    
    
    fill_gray = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    fill_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    for row_num, row_data in enumerate(data, 2):
        row_fill = fill_gray if ((row_num - 2) % 2 == 0) else fill_blue
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill = row_fill
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2


    if missing:
        start_row = len(data) + 3
        title_cell = ws.cell(row=start_row, column=1, value="Missing/No match:")
        title_cell.font = Font(bold=True, size=12)
        ws.cell(row=start_row, column=2, value=missing)
    
    filename = f"{name}.xlsx"
    
    wb.save(filename)
    print(f"Excel file created successfully: {filename}")
    
    return filename
